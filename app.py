from flask import Flask, render_template, request, send_file, jsonify, session
import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException
import datetime
import os
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


app = Flask(__name__)
app.secret_key = os.urandom(24)  # Needed for session management

def fetch_driver_expiry_date(driver_licence_no, driver):
    try:
        full_driver_licence_no = str(driver_licence_no)  
        search_driver_licence_no = full_driver_licence_no  # Use full badge number directly

        print(f"Searching for driver with full badge number: {search_driver_licence_no}")

        # Navigate to the driver licence search page
        driver.get("https://tph.tfl.gov.uk/TfL/SearchDriverLicence.page?org.apache.shale.dialog.DIALOG_NAME=TPHDriverLicence&Param=lg2.TPHDriverLicence&menuId=6")
        
        # Wait for the input field to be available and then fill it
        WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.ID, "searchdriverlicenceform:DriverLicenceNo")))

        # Clear and input the full driver licence number
        driver_licence_input = driver.find_element(By.ID, "searchdriverlicenceform:DriverLicenceNo")
        driver_licence_input.clear()
        driver_licence_input.send_keys(search_driver_licence_no)

        # Wait for the submit button to be clickable and click it
        submit_button = WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.ID, "searchdriverlicenceform:_id189"))
        )
        submit_button.click()

        
        # Wait for the next page to load and check for results
        WebDriverWait(driver, 20).until(EC.url_contains("https://tph.tfl.gov.uk/TfL/lg2/TPHLicensing/pubregsearch/Driver/SearchDriverLicence.page"))

        # Wait for the driver results table to appear
        try:
            error_message = driver.find_elements(By.ID, "validation")
            if error_message:
                return "Revoked & suspended to work"
            # This is the tbody containing the results
            WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.ID, "_id177:driverResults:tbody_element")))
        except TimeoutException:
            return "Error: Could not find the driver results."

        # Now that the table is present, let's extract the expiry date
        rows = driver.find_elements(By.CSS_SELECTOR, "#_id177\\:driverResults\\:tbody_element tr")

        for row in rows:
            columns = row.find_elements(By.TAG_NAME, "td")
            if len(columns) >= 3:  # If the row has at least 3 columns
                expiry_date = columns[2].text.strip()  # The expiry date is in the 3rd column
                return expiry_date

        return "Expiry Date Not Found"

    except TimeoutException as e:
        print(f"TimeoutException: The page took too long to load: {str(e)}")
        return "Error: Timeout"
    
    except NoSuchElementException as e:
        print(f"NoSuchElementException: Element not found: {str(e)}")
        return "Error: Element not found"
    
    except Exception as e:
        print(f"Unexpected error: {str(e)}")
        return "Revoked & suspended to work"

# Function to fetch expiry date for vehicles (using VRM)
def fetch_expiry_date(vrm, driver):
    try:
        driver.get("https://tph.tfl.gov.uk/TfL/SearchVehicleLicence.page?org.apache.shale.dialog.DIALOG_NAME=TPHVehicleLicence&Param=lg2.TPHVehicleLicence&menuId=7")
        WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.ID, "searchvehiclelicenceform:VehicleVRM")))
        vrm_input = driver.find_element(By.ID, "searchvehiclelicenceform:VehicleVRM")
        vrm_input.clear()
        vrm_input.send_keys(vrm)
        submit_button = driver.find_element(By.ID, "searchvehiclelicenceform:_id187")
        submit_button.click()
        WebDriverWait(driver, 20).until(EC.url_contains("https://tph.tfl.gov.uk/TfL/lg2/TPHLicensing/pubregsearch/Vehicle/SearchVehicleLicence.page"))
        
        error_message = driver.find_elements(By.ID, "validation")
        if error_message:
            return "Revoked & suspended to work"

        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "_id173")))

        paragraphs = driver.find_elements(By.TAG_NAME, "p")
        for para in paragraphs:
            if "Licence Expiry Date:" in para.text:
                expiry_date = para.text.split("Licence Expiry Date:")[1].strip()
                return expiry_date
        return "Expiry Date Not Found"
    except Exception as e:
        return "Error fetching expiry date"
    

def delete_old_files(is_drivers):
    # Define directories for driver and vehicle
    driver_dir = 'downloadedFiles/driver'
    vehicle_dir = 'downloadedFiles/vehicle'

    # Delete files based on whether it's for drivers or vehicles
    if is_drivers:
        # Delete all files in the driver folder
        for filename in os.listdir(driver_dir):
            file_path = os.path.join(driver_dir, filename)
            if os.path.isfile(file_path):
                os.remove(file_path)
        print("Deleted old files in 'downloadedFiles/driver'")
    else:
        # Delete all files in the vehicle folder
        for filename in os.listdir(vehicle_dir):
            file_path = os.path.join(vehicle_dir, filename)
            if os.path.isfile(file_path):
                os.remove(file_path)
        print("Deleted old files in 'downloadedFiles/vehicle'")



def color_rows(output_path):
    """
    Applies red coloring to rows containing specific keywords in the Comments column.
    """
    try:
        # Load the Excel file
        wb = load_workbook(output_path)
        ws = wb.active  # Use the active worksheet

        # Define the red fill style
        red_fill = PatternFill(start_color="FF8488", end_color="FF8488", fill_type="solid")

        # Find the "Comments" column index dynamically
        comments_col = None
        for idx, cell in enumerate(ws[1], start=1):  # Assuming the first row contains headers
            if cell.value and "comments" in cell.value.lower():
                comments_col = idx
                break

        if comments_col is None:
            raise ValueError("'Comments' column not found in the Excel sheet.")

        # Debugging: Print comments column index
        print(f"'Comments' column found at index: {comments_col}")

        # Define keywords to look for in the Comments column
        keywords = ["revoked & suspended to work", "revoked & expired license"]

        # Iterate through each row and apply the red fill to rows containing the keywords
        rows_colored = 0
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column), start=2):
            cell_value = row[comments_col - 1].value  # Get the value in the Comments column
            # Debugging: Print the exact cell value being checked
            print(f"Row {row_idx}: Comments Value - {cell_value}")
            if cell_value and any(keyword in str(cell_value).lower() for keyword in keywords):
                # Apply red fill to all cells in the row
                print(f"Coloring row {row_idx} for value: {cell_value}")
                for cell in row:
                    cell.fill = red_fill
                rows_colored += 1

        # Save the updated file
        wb.save(output_path)
        wb.close()

        print(f"Row coloring applied successfully to {rows_colored} row(s) in file: {output_path}")
    except Exception as e:
        print(f"Error while applying row coloring: {e}")


def process_xlsx(file, driver, is_drivers=False):
    try:
        # Read the Excel file into a pandas DataFrame
        df = pd.read_excel(file)

        # Debugging: Print column names for verification
        print("Columns in uploaded file:", df.columns)

        # Normalize column names for consistent access
        df.columns = df.columns.str.strip().str.lower()

        # Determine the correct expiry column based on file type
        if is_drivers:
            expiry_column_name = "badge expires"  # For driver files
        else:
            expiry_column_name = "plate expires"  # For vehicle files

        # Check if the required expiry column exists
        if expiry_column_name not in df.columns:
            raise ValueError(f"'{expiry_column_name}' column is missing in the uploaded file.")

        # Add new columns for "New Expiry Date" and "Comments" if they don't exist
        if "new expiry date" not in df.columns:
            df.insert(df.columns.get_loc(expiry_column_name) + 1, "new expiry date", "")  # Add after expiry column
        if "comments" not in df.columns:
            df.insert(df.columns.get_loc("new expiry date") + 1, "comments", "")  # Add after new expiry date

        # Iterate through each row in the DataFrame
        for index, row in df.iterrows():
            if is_drivers:
                # Process driver data (Badge Number)
                badge_number = str(row.iloc[3]).strip()  # Assuming Badge Number is in Column D
                if len(badge_number) <= 4:
                    continue  # Skip invalid badge numbers

                # Fetch the expiry date using the badge number
                expiry_date = fetch_driver_expiry_date(badge_number[:-4], driver)
            else:
                # Process vehicle data (VRM)
                vrm = str(row.iloc[3]).replace(" ", "").strip()  # Assuming VRM is in Column E
                if not vrm:
                    continue  # Skip invalid VRMs

                # Fetch the expiry date using the VRM
                expiry_date = fetch_expiry_date(vrm, driver)

            # Handle results based on fetched expiry date
            if "Error" in expiry_date or expiry_date == "Revoked & suspended to work":
                # Leave the new expiry date empty, retain original date, and add a comment
                df.at[index, "new expiry date"] = ""
                df.at[index, "comments"] = "Revoked & expired license"
            else:
                # Convert and store the fetched expiry date
                df.at[index, "new expiry date"] = convert_date(expiry_date)
                df.at[index, "comments"] = ""  # Clear comments for valid entries

            # Ensure original expiry date is properly converted for consistency
            original_expiry = row[expiry_column_name]
            df.at[index, expiry_column_name] = convert_date(original_expiry)

        # Generate output filename
        current_time = datetime.now().strftime('%Y%m%d%H%M%S')
        if is_drivers:
            output_filename = f"DriverLicense-{current_time}.xlsx"
            output_path = os.path.join('downloadedFiles/driver', output_filename)
        else:
            output_filename = f"VehicleLicense-{current_time}.xlsx"
            output_path = os.path.join('downloadedFiles/vehicle', output_filename)

        # Save the DataFrame to Excel
        df.to_excel(output_path, index=False)

        # Apply row coloring for 'Revoked & suspended to work'
        color_rows(output_path)

        return output_filename
    except Exception as e:
        print(f"Error during processing: {e}")
        return None


def convert_date(date_value):
    """
    Converts a given date value to 'dd/mm/yyyy' format.
    Handles both string and numeric (Excel serial) date values.
    """
    try:
        # Handle numeric (Excel serial) dates
        if isinstance(date_value, (int, float)):
            # Excel serial date starts from 1900-01-01
            converted_date = datetime.fromordinal(datetime(1900, 1, 1).toordinal() + int(date_value) - 2)
            return converted_date.strftime('%d/%m/%Y')
        
        # Handle string dates
        elif isinstance(date_value, str):
            try:
                # Attempt to parse the date from string
                converted_date = datetime.strptime(date_value, '%d/%m/%Y')
                return converted_date.strftime('%d/%m/%Y')
            except ValueError:
                # Handle alternative formats if needed
                return date_value  # Return as-is if parsing fails

        # Return empty if no valid date is found
        return ""
    except Exception as e:
        print(f"Error converting date: {e}")
        return ""









@app.route('/')
def index():
    is_drivers = request.args.get('is_drivers', 'false')  # Default to 'false' if not set
    return render_template('index.html', is_drivers=is_drivers)

@app.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return jsonify({"error": "No file uploaded"}), 400
    file = request.files['file']
    is_drivers = request.form.get('is_drivers', 'false') == 'true'  # Get the 'is_drivers' value from the form

    # Set up Selenium WebDriver (headless)
    options = webdriver.ChromeOptions()
    # options.add_argument("--headless")
    driver = webdriver.Chrome(options=options)

    # Process the file in a separate thread or background task
    output_filename = process_xlsx(file, driver, is_drivers)
    driver.quit()

    return jsonify({"download_url": f"/download/{output_filename}?is_drivers={str(is_drivers).lower()}"}), 200


@app.route('/progress')
def get_progress():
    # Send the current progress
    total = session.get('total_rows', 0)
    processed = session.get('processed_rows', 0)
    return jsonify({"processed": processed, "total": total})

@app.route('/download/<filename>')
def download_file(filename):
    # Get the 'is_drivers' query parameter (default to False if not provided)
    is_drivers = request.args.get('is_drivers', 'false') == 'true'
    
    # Determine the subdirectory based on 'is_drivers'
    if is_drivers:
        file_path = os.path.join('downloadedFiles', 'driver', filename)
    else:
        file_path = os.path.join('downloadedFiles', 'vehicle', filename)
    
    # Check if the file exists
    if not os.path.exists(file_path):
        return jsonify({"error": "File not found"}), 404
    
    # Send the file as a download response
    return send_file(file_path, as_attachment=True)

if __name__ == "__main__":
    app.run(debug=True)